from openpyxl import workbook, load_workbook

import excel
import json

operacje = ['buy','sell']
waluty = ['pln', 'eur', 'usd', 'gbp']


razemUSD = json.dollarRate * excel.currentUSD
razemPLN = excel.currentPLN
razemEUR= json.euroRate * excel.currentEUR
razemGBP = json.poundRate * excel.currentGBP

razem = razemGBP + razemEUR + razemPLN + razemUSD

def checkAktualnyStan():
    worksheet = excel.workbook['Aktualny Stan']
    currentPLN = worksheet['B' + str(excel.findNextEmptyRow() - 1)].value
    currentEUR = worksheet['C' + str(excel.findNextEmptyRow() - 1)].value
    currentUSD = worksheet['D' + str(excel.findNextEmptyRow() - 1)].value
    currentGBP = worksheet['E' + str(excel.findNextEmptyRow() - 1)].value
    worksheet['F' + str(excel.findNextEmptyRow())] = razem
    excel.closeWorkbook()
    return currentPLN, currentEUR, currentUSD, currentGBP

def writeAktualnyStan():
    worksheet = excel.workbook['Aktualny Stan']
    #zapisac zmienne po operacjach
    worksheet['F'+str(excel.findNextEmptyRow())] = razem
    excel.closeWorkbook()




