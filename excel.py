from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

#Workbook - excel file
#Worksheet - current sheet in the excel file

workbook = load_workbook('waluty.xlsx')
worksheet = workbook.active

def closeWorkbook():
    workbook.save('waluty.xlsx')

currentPLN = worksheet['B2'].value
currentEUR = worksheet['B3'].value
currentUSD = worksheet['B4'].value
currentGBP = worksheet['B5'].value


#Loop through rows to find the first empty one
def findNextEmptyRow():
    for cell in worksheet["A"]:
        if cell.value is None:
            return cell.row
